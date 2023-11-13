import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def aggiungi_spesa(data, tipo, categoria, importo, note):
    file_path = 'Expense_tracker.xlsx'
    
    # Crea un nuovo foglio di lavoro se il file non esiste
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Data', 'Tipo', 'Categoria', 'Importo', 'Note'])
        workbook.save(file_path)

    # Apri il foglio di lavoro esistente
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Aggiungi la nuova spesa
    sheet.append([data, tipo, categoria, importo, note])

    # Salva il foglio di lavoro
    workbook.save(file_path)

def main():
    print("Expense Tracker")

    categorie_disponibili = ['casa', 'gasolio', 'autostrada', 'bar', 'pasti fuori']
    tipo_di_spesa = ['in', 'out']

    while True:
        data = input("Inserisci la data (formato GG/MM/AAAA): ")
        
        print("tipo di spesa. in o out?")
        
        tipo = input("Che tipo di spesa è?: ").lower()
        while tipo not in tipo_di_spesa:
            print("Tipo di spesa non valido. Scegli tra le categorie suggerite.")
            tipo = input("Che tipo di spesa è?:")

        print("Categorie suggerite:", ", ".join(categorie_disponibili))
       
        categoria = input("Per quale categoria? ").lower()  # Converti la categoria in minuscolo
        while categoria not in categorie_disponibili:
            print("Categoria non valida. Scegli tra le categorie suggerite o aggiungine una nuova.")
            categoria = input("Per quale categoria? ").lower()

        importo = float(input("Inserisci l'importo della spesa: "))

        note = ""
        scelta_note = input("Vuoi aggiungere delle note alla spesa? (s/n): ")
        if scelta_note.lower() == 's':
            note = input("Inserisci la nota: ")
        
        aggiungi_spesa(data, tipo, categoria, importo, note)

        scelta = input("Vuoi aggiungere un'altra spesa? (s/n): ")
        if scelta.lower() != 's':
            break

if __name__ == "__main__":
    main()
